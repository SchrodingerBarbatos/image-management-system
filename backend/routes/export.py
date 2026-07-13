import os, sys, uuid, zipfile, datetime, threading, logging, traceback, re, json, io
from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook, Workbook
from models import session, Image, ExportTask, ScanRoot, BarcodeSetting, ImageVersion
from config import UPLOAD_DIR, ZIP_CLEANUP_HOURS, XLSX_TTL_HOURS


def _xlsx_dir():
    d = os.path.join(UPLOAD_DIR, 'xlsx')
    os.makedirs(d, exist_ok=True)
    return d


def _zip_dir():
    d = os.path.join(UPLOAD_DIR, 'zips')
    os.makedirs(d, exist_ok=True)
    return d

export_bp = Blueprint('export', __name__)
_log = logging.getLogger(__name__)

_export_lock = threading.Lock()

# Maximum number of parameters per IN clause — keeps SQLite happy
_IN_CHUNK_SIZE = 500


def _safe_remove(path):
    """Remove a file, silently ignoring errors (e.g. file not found)."""
    try:
        os.remove(path)
    except OSError:
        pass


def _chunked_in_query(column, values, query_base, chunk_size=_IN_CHUNK_SIZE):
    """Execute a query with a potentially large IN clause by splitting into chunks.
    Returns the concatenated results of all chunks."""
    if not values:
        return []
    all_rows = []
    value_list = list(values)
    for i in range(0, len(value_list), chunk_size):
        chunk = value_list[i:i + chunk_size]
        rows = query_base.filter(column.in_(chunk)).all()
        all_rows.extend(rows)
    return all_rows


def filter_to_single_version(imgs, barcodes, session):
    """Return subset of imgs containing only one version per (barcode, image_type).

    Priority chain (first match wins):
    1. BarcodeSetting.default_{main,detail}_ctime — user-chosen default per barcode
    2. ImageVersion.is_latest — most recent version detected by scanner
    3. Pass-through — if no version data exists for a barcode, keep all its images

    Relies on ImageVersion and BarcodeSetting tables for version resolution.
    When a barcode has neither, images are kept as-is (fallback to "all images").
    """
    # Chunked query for BarcodeSetting
    settings = _chunked_in_query(
        BarcodeSetting.barcode, barcodes,
        session.query(
            BarcodeSetting.barcode, BarcodeSetting.default_main_ctime, BarcodeSetting.default_detail_ctime
        ),
    )
    # Chunked query for ImageVersion
    latest_versions = _chunked_in_query(
        ImageVersion.barcode, barcodes,
        session.query(
            ImageVersion.barcode, ImageVersion.image_type, ImageVersion.folder_ctime
        ).filter(ImageVersion.is_latest == True),
    )
    allowed = {}  # {barcode: {image_type: folder_ctime}}
    for v in latest_versions:
        allowed.setdefault(v.barcode, {})[v.image_type] = v.folder_ctime
    for s in settings:
        if s.default_main_ctime:
            allowed.setdefault(s.barcode, {})['main'] = s.default_main_ctime
        if s.default_detail_ctime:
            allowed.setdefault(s.barcode, {})['detail'] = s.default_detail_ctime
    filtered = []
    for img in imgs:
        type_map = allowed.get(img.barcode)
        if type_map:
            allowed_ctime = type_map.get(img.image_type)
            if allowed_ctime and img.folder_ctime != allowed_ctime:
                continue
        filtered.append(img)
    return filtered

def _compute_barcode_counts(imgs, all_barcodes=None, export_type='all'):
    """Compute per-barcode main/detail match counts for the export report.

    export_type:
    - 'all': real main/detail counts from imgs
    - 'main': only main counts (detail forced 0)
    - 'detail': detail counts; barcodes that only have main (fallback) report
      those as detail so the detail Excel matches ZIP contents
    """
    counts = {}
    if all_barcodes:
        for b in all_barcodes:
            counts[b] = {'main': 0, 'detail': 0}

    if export_type == 'detail':
        grouped = {}
        for img in imgs:
            grouped.setdefault(img.barcode, {'main': 0, 'detail': 0})
            if img.image_type in ('main', 'detail'):
                grouped[img.barcode][img.image_type] += 1
        for bc, g in grouped.items():
            counts.setdefault(bc, {'main': 0, 'detail': 0})
            if g['detail'] > 0:
                counts[bc]['detail'] = g['detail']
                counts[bc]['main'] = 0
            else:
                # main-as-detail fallback: report under detail so report matches ZIP
                counts[bc]['detail'] = g['main']
                counts[bc]['main'] = 0
        return counts

    for img in imgs:
        counts.setdefault(img.barcode, {'main': 0, 'detail': 0})
        if img.image_type in ('main', 'detail'):
            counts[img.barcode][img.image_type] += 1

    if export_type == 'main':
        for bc in counts:
            counts[bc]['detail'] = 0
    return counts

def _col_letter(idx):
    """Convert 0-based column index to Excel column letter(s). 0->A, 25->Z, 26->AA."""
    result = ''
    while idx >= 0:
        result = chr(65 + (idx % 26)) + result
        idx = idx // 26 - 1
    return result


def _col_letter_to_idx(s):
    """Convert Excel column letter(s) to 0-based index. A->0, Z->25, AA->26."""
    s = s.strip().upper()
    if not s or not s.isalpha():
        raise ValueError('invalid column letter')
    idx = 0
    for c in s:
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx - 1

def _plan_zip_entries(img_data, flat, export_type='all'):
    """Compute the (file_path, arcname, scan_root_id) entries a ZIP export will contain.

    Pure function over img_data — no DB or filesystem access, so routes can
    call it to report an accurate entry count before the build thread starts.

    img_data rows are either:
      (file_path, barcode, image_type, sequence, ext)
    or (file_path, barcode, image_type, sequence, ext, scan_root_id)

    export_type semantics:
    - 'main': only main images, named as main
    - 'detail': detail images; barcodes with no detail get ALL their main
      images renamed as detail (fallback)
    - 'all': main as main, detail as detail, no fallback

    Returns (entries, fallback_barcodes) where each entry is
    (file_path, arcname, scan_root_id|None).
    """
    def _unpack(row):
        if len(row) >= 6:
            return row[0], row[1], row[2], row[3], row[4], row[5]
        return row[0], row[1], row[2], row[3], row[4], None

    entries = []
    fallback_barcodes = []

    if export_type == 'main':
        for row in img_data:
            file_path, barcode, image_type, sequence, ext, root_id = _unpack(row)
            if image_type != 'main':
                continue
            display_name = f"{barcode}_{sequence}.{ext}"
            entries.append((file_path, display_name if flat else f"主图/{display_name}", root_id))

    elif export_type == 'detail':
        grouped = {}
        for row in img_data:
            fp, bc, it, seq, ex, root_id = _unpack(row)
            grouped.setdefault(bc, {'main': [], 'detail': []})
            if it in ('main', 'detail'):
                grouped[bc][it].append((fp, seq, ex, root_id))
        for barcode, groups in grouped.items():
            if groups['detail']:
                items = groups['detail']
            else:
                items = groups['main']
                if items:
                    fallback_barcodes.append(barcode)
            for file_path, sequence, ext, root_id in items:
                display_name = f"{barcode}_详情图_{sequence}.{ext}"
                entries.append((file_path, display_name if flat else f"详情图/{display_name}", root_id))

    else:
        for row in img_data:
            file_path, barcode, image_type, sequence, ext, root_id = _unpack(row)
            if image_type == 'main':
                display_name = f"{barcode}_{sequence}.{ext}"
                arcname = display_name if flat else f"主图/{display_name}"
            else:
                display_name = f"{barcode}_详情图_{sequence}.{ext}"
                arcname = display_name if flat else f"详情图/{display_name}"
            entries.append((file_path, arcname, root_id))

    return entries, fallback_barcodes


def _build_zip(task_id, img_data, flat, export_type='all'):
    """Build ZIP file in a background thread, updating task progress.

    img_data: list of (file_path, barcode, image_type, sequence, ext) tuples — plain
    data so the thread doesn't need access to the request-scoped session.
    export_type: 'main', 'detail', or 'all'. Controls naming and fallback
    (see _plan_zip_entries).
    """
    from models import session as sess, ExportTask
    try:
        task = sess.get(ExportTask, task_id)
        if not task:
            _log.warning("_build_zip: task %s not found (may have been deleted)", task_id)
            return
        entries, fallback_barcodes = _plan_zip_entries(img_data, flat, export_type)
        total = len(entries)
        if total == 0:
            task.status = 'done'
            task.total_images = 0
            task.progress = 0
            zip_path = os.path.join(_zip_dir(), f'export_{task_id}.zip')
            task.zip_path = zip_path
            with zipfile.ZipFile(zip_path, 'w'):
                pass
            sess.commit()
            return
        task.total_images = total
        task.progress = 0
        sess.commit()

        zip_name = f'export_{task_id}.zip'
        zip_path = os.path.join(_zip_dir(), zip_name)
        task.zip_path = zip_path
        sess.commit()

        written = 0
        progress = 0
        actual_counts = {}  # barcode -> {main, detail} based on successful writes

        # Path confinement: file must live under ITS OWN enabled ScanRoot
        # (not any root). Uses realpath/commonpath via is_path_under_root.
        from models import ScanRoot
        from routes._utils import is_path_under_root
        root_map = {
            r.id: r for r in sess.query(ScanRoot).all()
        }

        def _allowed_for_root(fp: str, root_id) -> bool:
            if root_id is None:
                return False
            root = root_map.get(root_id)
            if not root or not root.enabled:
                return False
            ok, _ = is_path_under_root(fp, root.path)
            return ok

        # Map arcname -> barcode for actual count attribution
        # entries are (file_path, arcname, scan_root_id); barcode is embedded in arcname
        # Prefer tracking from original img_data order via planned entries
        entry_barcodes = []
        # Rebuild barcode list parallel to entries by re-planning with tags
        for row in img_data:
            if len(row) >= 6:
                fp, bc, it, seq, ex, rid = row[0], row[1], row[2], row[3], row[4], row[5]
            else:
                fp, bc, it, seq, ex, rid = row[0], row[1], row[2], row[3], row[4], None
            # Will re-derive from entries + export_type below

        # Build (file_path, arcname, root_id, barcode, report_type) for counting
        typed_entries = []
        for entry in entries:
            if len(entry) == 3:
                file_path, arcname, root_id = entry
            else:
                file_path, arcname = entry[0], entry[1]
                root_id = None
            # Find matching img_data row by file_path
            barcode = None
            report_type = 'main' if '/主图/' in arcname.replace('\\', '/') or (
                export_type == 'main' or (export_type == 'all' and '_详情图_' not in arcname)
            ) else 'detail'
            if export_type == 'detail' or '_详情图_' in arcname:
                report_type = 'detail'
            elif export_type == 'main' or (export_type == 'all' and '_详情图_' not in os.path.basename(arcname)):
                report_type = 'main'
            for row in img_data:
                if row[0] == file_path:
                    barcode = row[1]
                    # For detail export fallback, count under detail
                    if export_type == 'detail':
                        report_type = 'detail'
                    elif export_type == 'main':
                        report_type = 'main'
                    else:
                        report_type = row[2] if row[2] in ('main', 'detail') else 'main'
                    break
            typed_entries.append((file_path, arcname, root_id, barcode, report_type))

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_STORED) as zf:
            for file_path, arcname, root_id, barcode, report_type in typed_entries:
                if os.path.exists(file_path) and _allowed_for_root(file_path, root_id):
                    zf.write(file_path, arcname)
                    written += 1
                    if barcode:
                        actual_counts.setdefault(barcode, {'main': 0, 'detail': 0})
                        if report_type in ('main', 'detail'):
                            actual_counts[barcode][report_type] += 1
                elif os.path.exists(file_path):
                    _log.warning(
                        "_build_zip: skip path not under own enabled root "
                        "(root_id=%s): %s", root_id, file_path,
                    )
                progress += 1
                if progress % 100 == 0:
                    task.progress = progress
                    sess.commit()

        if fallback_barcodes:
            _log.info("_build_zip task %s: %d barcodes used main images as detail fallback",
                      task_id, len(fallback_barcodes))

        # Always finalize progress to planned total so UI never sticks < 100%
        task.progress = total
        skipped = total - written
        if written == 0:
            _log.warning("_build_zip task %s: all %d files missing/skipped", task_id, total)
            task.status = 'failed'
            task.error_message = '所有匹配的图片文件均不存在或不在所属扫描目录下'
        elif written < total:
            task.status = 'partial_failed'
            task.error_message = f'实际写入 {written} / 计划 {total}（跳过 {skipped}）'
            _log.info("_build_zip task %s: partial_failed wrote %d/%d", task_id, written, total)
        else:
            task.status = 'done'

        # Persist structured payload: barcodes = actual written counts; stats separate
        # Preserve unmatched barcodes from original payload (zeros) for report completeness
        try:
            prev = json.loads(task.barcode_data) if task.barcode_data else {}
        except (json.JSONDecodeError, TypeError):
            prev = {}
        prev_barcodes, _ = _parse_export_payload(task.barcode_data) if task.barcode_data else ({}, {})
        if prev_barcodes is None:
            prev_barcodes = {}
        # Start from planned zeros, overlay actual written
        final_barcodes = {
            bc: {'main': 0, 'detail': 0} for bc in prev_barcodes
        }
        for bc, c in actual_counts.items():
            final_barcodes.setdefault(bc, {'main': 0, 'detail': 0})
            final_barcodes[bc]['main'] = c.get('main', 0)
            final_barcodes[bc]['detail'] = c.get('detail', 0)
        task.barcode_data = json.dumps({
            'barcodes': final_barcodes,
            'stats': {
                'planned_count': total,
                'written_count': written,
                'skipped_count': skipped,
            },
        }, ensure_ascii=False)
        sess.commit()
    except Exception as e:
        _log.error("_build_zip task %s failed:\n%s", task_id, traceback.format_exc())
        try:
            task = sess.get(ExportTask, task_id)
            if task:
                task.status = 'failed'
                if getattr(sys, 'frozen', False):
                    task.error_message = f"导出失败: {e}"
                else:
                    task.error_message = traceback.format_exc()
                sess.commit()
        except Exception:
            pass
    finally:
        sess.remove()


@export_bp.route('/export/excel', methods=['POST'])
def upload_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'file required'}), 400

    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': '只允许 .xlsx 文件'}), 400

    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({'error': '文件大小不能超过 10MB'}), 400

    upload_id = uuid.uuid4().hex[:12]
    upload_path = os.path.join(_xlsx_dir(), f'{upload_id}.xlsx')
    try:
        file.save(upload_path)
        # Parse only — keep file for ZIP generation; TTL cleanup covers abandon.
        data, error = _parse_excel(upload_path)
        if error:
            _safe_remove(upload_path)
            return jsonify({'error': error}), 400
        return jsonify({**data, 'upload_id': upload_id})
    except Exception as e:
        _safe_remove(upload_path)
        _log.exception("upload_excel failed: %s", e)
        return jsonify({'error': '上传处理失败'}), 500


def _parse_excel(upload_path):
    """Parse an xlsx file and return (data_dict, error_string).

    Returns ({columns, sheets, sheet_columns}, None) on success,
    or (None, error_message) on failure.  Caller is responsible for
    cleaning up upload_path on error.
    """
    wb = None
    try:
        wb = load_workbook(upload_path, read_only=True)
    except Exception:
        return None, '无法解析 Excel 文件，请确认文件格式正确'

    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            return None, 'Excel 文件中没有工作表'

        sheet_columns = {}
        for sname in sheet_names:
            ws = wb[sname]
            row = next(ws.iter_rows(min_row=1, max_row=1), None)
            if row is None:
                sheet_columns[sname] = []
            else:
                headers = [str(cell.value) for cell in row]
                if not headers or all(h == 'None' for h in headers):
                    sheet_columns[sname] = []
                else:
                    sheet_columns[sname] = [f'{_col_letter(i)}-{h}' for i, h in enumerate(headers)]
    finally:
        wb.close()

    if not sheet_columns or all(v == [] for v in sheet_columns.values()):
        return None, '所有工作表的表头均为空，请检查 Excel 文件'

    return {
        'columns': sheet_columns.get(sheet_names[0], []),
        'sheets': sheet_names,
        'sheet_columns': sheet_columns,
    }, None


@export_bp.route('/export/zip', methods=['POST'])
def generate_zip():
    data = request.json
    barcode_col = data.get('barcode_column', '')
    image_type = data.get('image_type', '')
    upload_id = data.get('upload_id', '')
    sheet_name = data.get('sheet_name', '')
    selected = data.get('selected_barcodes')
    flat = data.get('flat', False)

    if not upload_id:
        return jsonify({'error': 'upload_id is required'}), 400

    # Validate upload_id format
    if not re.match(r'^[0-9a-f]{12}$', upload_id):
        return jsonify({'error': 'invalid upload_id'}), 400

    # Parse barcode column letter
    col_letter = barcode_col.split('-')[0] if '-' in barcode_col else 'A'
    try:
        col_idx = _col_letter_to_idx(col_letter)
    except ValueError:
        return jsonify({'error': f'invalid column letter: {col_letter}'}), 400

    # Read barcodes from Excel (xlsx subdir preferred, legacy flat path fallback)
    upload_path = os.path.join(_xlsx_dir(), f'{upload_id}.xlsx')
    if not os.path.isfile(upload_path):
        upload_path = os.path.join(UPLOAD_DIR, f'{upload_id}.xlsx')
    # Resolve real path and ensure it stays within UPLOAD_DIR tree
    real_upload = os.path.realpath(upload_path)
    real_upload_dir = os.path.realpath(UPLOAD_DIR)
    try:
        if os.path.commonpath([real_upload, real_upload_dir]) != real_upload_dir:
            return jsonify({'error': 'invalid upload_id'}), 400
    except ValueError:
        return jsonify({'error': 'invalid upload_id'}), 400
    if not os.path.isfile(real_upload):
        return jsonify({'error': 'upload file not found'}), 404

    try:
        wb = load_workbook(real_upload, read_only=True)
    except Exception:
        _safe_remove(real_upload)
        return jsonify({'error': '无法解析 Excel 文件'}), 400

    try:
        if sheet_name and sheet_name not in wb.sheetnames:
            return jsonify({'error': f'工作表 "{sheet_name}" 不存在'}), 400

        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        # Get header row to validate column index
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        if col_idx >= len(header_row):
            return jsonify({'error': f'列索引 {col_letter}({col_idx}) 超出表头范围(共 {len(header_row)} 列)'}), 400

        barcodes_raw = []
        for row in ws.iter_rows(min_row=2):
            if col_idx >= len(row):
                continue
            val = str(row[col_idx].value).strip() if row[col_idx].value else ''
            if val:
                barcodes_raw.append(val)
    finally:
        # Close workbook only — keep source xlsx until ExportTask is committed
        # (or TTL cleanup). Deleting here would break 409/DB-failure retries.
        wb.close()

    if selected:
        selected_set = set(selected)
        barcodes_raw = [b for b in barcodes_raw if b in selected_set]

    if not barcodes_raw:
        return jsonify({'error': 'Excel 中未找到任何条码数据'}), 400

    # Deduplicate barcodes while preserving order (for consistent match semantics)
    barcodes = list(dict.fromkeys(barcodes_raw))

    # Find matching images — chunked IN query to avoid oversized SQL
    q_base = session.query(Image).filter(Image.confirmed == True).join(
        ScanRoot, Image.scan_root_id == ScanRoot.id
    ).filter(ScanRoot.enabled == True)
    # No type filter for 'all' or 'detail' — detail exports need main images
    # available as the fallback source (see _plan_zip_entries)
    if image_type and image_type not in ('all', 'detail'):
        q_base = q_base.filter(Image.image_type == image_type)

    imgs = []
    for i in range(0, len(barcodes), _IN_CHUNK_SIZE):
        chunk = barcodes[i:i + _IN_CHUNK_SIZE]
        rows = q_base.filter(Image.barcode.in_(chunk)).all()
        imgs.extend(rows)

    # Filter to single version: user-chosen default, or latest version as fallback
    imgs = filter_to_single_version(imgs, barcodes, session)
    matched_barcodes = set(img.barcode for img in imgs)
    excluded_barcodes = len(barcodes) - len(matched_barcodes)

    # Compute planned per-barcode counts (include all barcodes, even unmatched)
    planned_counts = _compute_barcode_counts(imgs, barcodes, export_type=image_type or 'all')
    payload = {
        'barcodes': planned_counts,
        'stats': {'planned_count': 0, 'written_count': 0, 'skipped_count': 0},
    }

    # Concurrency guard: check + create + commit must be atomic
    with _export_lock:
        running = session.query(ExportTask).filter(ExportTask.status == 'processing').first()
        if running:
            return jsonify({'error': '已有导出任务正在执行中，请等待完成'}), 409

        task = ExportTask(status='processing', barcode_data=json.dumps(payload, ensure_ascii=False))
        session.add(task)
        session.commit()
        # Source xlsx only deleted after task is durable — 409/DB failure keeps
        # the file for retry; TTL cleanup handles abandon.
        _safe_remove(real_upload)

    img_data = [
        (img.file_path, img.barcode, img.image_type, img.sequence, img.ext, img.scan_root_id)
        for img in imgs
    ]
    # Accurate entry count for the sync response — detail exports may rename
    # main images as fallback, so the ZIP entry count can differ from len(imgs)
    planned_entries, _ = _plan_zip_entries(img_data, flat, image_type or 'all')
    # Update planned_count now that we know the true entry count
    with _export_lock:
        t = session.get(ExportTask, task.id)
        if t:
            t.total_images = len(planned_entries)
            try:
                data = json.loads(t.barcode_data) if t.barcode_data else payload
            except (json.JSONDecodeError, TypeError):
                data = payload
            if isinstance(data, dict):
                data.setdefault('stats', {})['planned_count'] = len(planned_entries)
                t.barcode_data = json.dumps(data, ensure_ascii=False)
            session.commit()

    threading.Thread(target=_build_zip, args=(task.id, img_data, flat, image_type or 'all'), daemon=True).start()

    return jsonify({'task_id': task.id, 'total_images': len(planned_entries), 'total_barcodes': len(matched_barcodes), 'excluded_barcodes': excluded_barcodes})


def _parse_export_payload(raw):
    """Normalize barcode_data JSON into {barcodes, stats}.

    Supports:
    - new shape: {barcodes: {...}, stats: {...}}
    - legacy flat: {barcode: {main, detail}, __export_stats?: {...}}
    """
    try:
        data = json.loads(raw) if raw else {}
    except (json.JSONDecodeError, TypeError):
        return None, None
    if not isinstance(data, dict):
        return None, None
    if 'barcodes' in data and isinstance(data.get('barcodes'), dict):
        barcodes = data['barcodes']
        stats = data.get('stats') if isinstance(data.get('stats'), dict) else {}
        return barcodes, stats
    # Legacy: filter reserved keys
    barcodes = {
        k: v for k, v in data.items()
        if not str(k).startswith('_') and isinstance(v, dict)
    }
    stats = data.get('__export_stats') if isinstance(data.get('__export_stats'), dict) else {}
    return barcodes, stats


@export_bp.route('/export/progress/<int:task_id>')
def export_progress(task_id):
    task = session.get(ExportTask, task_id)
    if not task:
        return jsonify({'error': 'not found'}), 404
    written = None
    planned = task.total_images
    skipped = None
    barcodes, stats = _parse_export_payload(task.barcode_data)
    if stats:
        written = stats.get('written_count')
        planned = stats.get('planned_count', planned)
        skipped = stats.get('skipped_count')
    return jsonify({
        'status': task.status,
        'progress': task.progress,
        'total': task.total_images,
        'planned_count': planned,
        'written_count': written,
        'skipped_count': skipped,
        'error_message': task.error_message,
    })


@export_bp.route('/export/tasks')
def list_tasks():
    tasks = session.query(ExportTask).order_by(ExportTask.id.desc()).limit(30).all()
    result = []
    for t in tasks:
        file_available = bool(t.zip_path and os.path.exists(t.zip_path))
        result.append({
            'id': t.id, 'status': t.status, 'total_images': t.total_images,
            'created_at': t.created_at, 'file_available': file_available,
            'error_message': t.error_message,
            'has_detail': bool(t.barcode_data),
        })
    return jsonify(result)


@export_bp.route('/export/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    task = session.get(ExportTask, task_id)
    if not task:
        return jsonify({'error': 'not found'}), 404
    if request.args.get('force') != 'true' and task.status == 'processing':
        return jsonify({'error': 'cannot delete processing task'}), 409
    if task.zip_path and os.path.exists(task.zip_path):
        try:
            os.remove(task.zip_path)
        except OSError:
            pass
    session.delete(task)
    session.commit()
    return jsonify({'ok': True})


@export_bp.route('/export/download/<int:task_id>')
def download_zip(task_id):
    task = session.get(ExportTask, task_id)
    downloadable = {'done', 'partial_failed'}
    if not task or task.status not in downloadable:
        return jsonify({'error': 'not ready'}), 404
    if not task.zip_path:
        return jsonify({'error': 'file has been cleaned up'}), 404
    # Path must stay under the controlled zips directory
    real_zip = os.path.realpath(task.zip_path)
    real_zip_dir = os.path.realpath(_zip_dir())
    try:
        if os.path.commonpath([real_zip, real_zip_dir]) != real_zip_dir:
            _log.error("download_zip: path escape blocked for task %s: %s", task_id, task.zip_path)
            return jsonify({'error': 'invalid zip path'}), 404
    except ValueError:
        return jsonify({'error': 'invalid zip path'}), 404
    if not os.path.isfile(real_zip):
        return jsonify({'error': 'file has been cleaned up'}), 404
    return send_file(
        real_zip,
        as_attachment=True,
        download_name=f'export_{task_id}.zip',
        mimetype='application/zip',
        conditional=True,
    )


@export_bp.route('/export/tasks/<int:task_id>/detail')
def download_detail(task_id):
    task = session.get(ExportTask, task_id)
    if not task:
        return jsonify({'error': 'not found'}), 404
    if not task.barcode_data:
        return jsonify({'error': 'no barcode data available'}), 404

    barcode_counts, _stats = _parse_export_payload(task.barcode_data)
    if barcode_counts is None:
        return jsonify({'error': 'barcode data is corrupted'}), 500

    wb = Workbook()

    # Sheet 1: 导出详情（保留原有）
    ws_all = wb.active
    ws_all.title = '导出详情'
    ws_all.append(['条码', '匹配主图数量', '匹配详情图数量'])

    # Sheet 2: 主图匹配（新增）
    ws_main = wb.create_sheet('主图匹配')
    ws_main.append(['条码', '主图数量'])

    # Sheet 3: 详情图匹配（新增）
    ws_detail = wb.create_sheet('详情图匹配')
    ws_detail.append(['条码', '详情图数量'])

    # Write data to all sheets in single pass — only real barcodes
    for barcode, counts in barcode_counts.items():
        if not isinstance(counts, dict):
            continue
        main_count = counts.get('main', 0)
        detail_count = counts.get('detail', 0)
        ws_all.append([barcode, main_count, detail_count])
        ws_main.append([barcode, main_count])
        ws_detail.append([barcode, detail_count])

    # Auto-fit column widths for all sheets
    for ws in [ws_all, ws_main, ws_detail]:
        for col_idx, _ in enumerate(ws[1], start=1):
            max_width = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        val = str(cell.value)
                        width = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_width = max(max_width, width)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'export_detail_{task_id}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def reset_stale_processing():
    """Startup helper — mark ALL processing export tasks as failed
    so they don't permanently block new exports."""
    with _export_lock:
        stale = session.query(ExportTask).filter(
            ExportTask.status == 'processing',
        ).all()
        if stale:
            for task in stale:
                task.status = 'failed'
                task.error_message = '程序重启或导出进程异常中断'
            session.commit()
            _log.info("Reset %d stale processing export tasks to failed", len(stale))


def cleanup_old_exports():
    """Remove export tasks and their files older than ZIP_CLEANUP_HOURS.

    Also purge orphaned xlsx older than XLSX_TTL_HOURS and leftover zip files
    without a task row. Path-confined to UPLOAD_DIR tree. Individual delete
    failures are logged and skipped so one bad file cannot abort the run.
    """
    cutoff = datetime.datetime.now() - datetime.timedelta(hours=ZIP_CLEANUP_HOURS)
    with _export_lock:
        old_tasks = session.query(ExportTask).filter(
            ExportTask.created_at < cutoff.isoformat(),
            ExportTask.status != 'processing',
        ).all()
        for task in old_tasks:
            if task.zip_path and os.path.exists(task.zip_path):
                try:
                    # Confine to zips dir
                    real_zip = os.path.realpath(task.zip_path)
                    real_zip_dir = os.path.realpath(_zip_dir())
                    try:
                        if os.path.commonpath([real_zip, real_zip_dir]) == real_zip_dir:
                            os.remove(real_zip)
                    except ValueError:
                        pass
                except OSError as e:
                    _log.warning("cleanup: failed to remove zip %s: %s", task.zip_path, e)
            session.delete(task)
        if old_tasks:
            session.commit()
            _log.info("Cleaned up %d old export tasks", len(old_tasks))

    # TTL cleanup for source xlsx
    xlsx_cutoff = datetime.datetime.now() - datetime.timedelta(hours=XLSX_TTL_HOURS)
    for directory in (_xlsx_dir(), UPLOAD_DIR):
        try:
            real_dir = os.path.realpath(directory)
            for name in os.listdir(directory):
                if not name.endswith('.xlsx'):
                    continue
                path = os.path.join(directory, name)
                try:
                    real_path = os.path.realpath(path)
                    if os.path.commonpath([real_path, real_dir]) != real_dir:
                        continue
                    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(real_path))
                    if mtime < xlsx_cutoff:
                        os.remove(real_path)
                except (OSError, ValueError) as e:
                    _log.warning("cleanup: skip xlsx %s: %s", path, e)
        except OSError as e:
            _log.warning("cleanup: list dir %s failed: %s", directory, e)

    # Orphan zips under zips/ with no matching task
    try:
        known = {
            os.path.realpath(t.zip_path)
            for t in session.query(ExportTask).filter(ExportTask.zip_path != '').all()
            if t.zip_path
        }
        zip_dir = _zip_dir()
        real_zip_dir = os.path.realpath(zip_dir)
        for name in os.listdir(zip_dir):
            if not name.endswith('.zip'):
                continue
            path = os.path.realpath(os.path.join(zip_dir, name))
            try:
                if os.path.commonpath([path, real_zip_dir]) != real_zip_dir:
                    continue
            except ValueError:
                continue
            if path not in known:
                try:
                    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path))
                    if mtime < cutoff:
                        os.remove(path)
                except OSError as e:
                    _log.warning("cleanup: orphan zip %s: %s", path, e)
    except OSError as e:
        _log.warning("cleanup: orphan zip scan failed: %s", e)


_cleanup_stop = threading.Event()
_cleanup_thread = None
_cleanup_lock = threading.Lock()


def start_export_cleanup_loop(interval_seconds=900):
    """Start a daemon loop that runs cleanup_old_exports periodically.

    Single-process: guarded by _cleanup_lock so only one loop starts.
    Multi-process: each process has its own loop; cleanup is idempotent and
    path-confined, so concurrent runs are safe (best-effort).
    """
    global _cleanup_thread
    with _cleanup_lock:
        if _cleanup_thread is not None and _cleanup_thread.is_alive():
            return

        def _loop():
            while not _cleanup_stop.wait(interval_seconds):
                try:
                    cleanup_old_exports()
                except Exception:
                    _log.exception("periodic cleanup_old_exports failed")

        _cleanup_stop.clear()
        _cleanup_thread = threading.Thread(target=_loop, name='export-cleanup', daemon=True)
        _cleanup_thread.start()
        _log.info("Started export cleanup loop (interval=%ss)", interval_seconds)


def stop_export_cleanup_loop():
    _cleanup_stop.set()
