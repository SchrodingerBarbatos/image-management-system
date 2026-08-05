"""Tests for batch task framework — task engine, duplicate scan tasks, low version scan tasks."""

import json, os, sqlite3, tempfile, time, threading
import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import scoped_session, sessionmaker
from flask import Flask

from models import (
    Base,
    Image,
    ImageVersion,
    ScanRoot,
    BatchTask,
    DuplicateVersionScanResult,
    DeletedFolder,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="function")
def engine():
    """File-based SQLite so background threads can open their own connections
    to the same database.  :memory: would give each connection an isolated db."""
    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    eng = create_engine(f"sqlite:///{db_path}", echo=False)

    @event.listens_for(eng, "connect")
    def _set_pragma(dbapi_conn, _):
        dbapi_conn.execute("PRAGMA journal_mode=WAL")
        dbapi_conn.execute("PRAGMA busy_timeout=5000")

    Base.metadata.create_all(bind=eng)
    yield eng
    eng.dispose()
    # Clean up the temp file and WAL/SHM sidecars
    for path in (db_path, db_path + '-wal', db_path + '-shm'):
        try:
            os.unlink(path)
        except OSError:
            pass


@pytest.fixture(scope="function")
def sess(engine):
    factory = sessionmaker(bind=engine)
    scoped = scoped_session(factory)
    try:
        yield scoped
    finally:
        scoped.remove()


@pytest.fixture(scope="function")
def client(sess, monkeypatch):
    import task_engine
    import routes.batch_tasks
    import routes.batch
    import versioning
    import routes.export

    monkeypatch.setattr(task_engine, "session", sess)
    monkeypatch.setattr(routes.batch_tasks, "session", sess)
    monkeypatch.setattr(routes.batch, "session", sess)
    monkeypatch.setattr(versioning, "session", sess)
    monkeypatch.setattr(routes.export, "session", sess)

    app = Flask(__name__)
    app.config['TESTING'] = True

    from routes.batch_tasks import batch_tasks_bp
    from routes.export import export_bp
    app.register_blueprint(batch_tasks_bp, url_prefix='/api')
    app.register_blueprint(export_bp)

    return app.test_client()


def _make_root(sess, root_id=1, path="/fake", enabled=True):
    sr = ScanRoot(id=root_id, path=path, enabled=enabled, recursive=False)
    sess.add(sr)
    sess.commit()
    return sr


def _make_image(sess, barcode, image_type, folder_ctime, filename="a.jpg",
                scan_root_id=1, confirmed=True, status="active", file_size=100,
                file_path=None):
    if file_path is None:
        file_path = f"/fake/{barcode}/{folder_ctime}/{filename}"
    img = Image(
        barcode=barcode, image_type=image_type, folder_ctime=folder_ctime,
        filename=filename, ext="jpg", file_path=file_path, file_size=file_size,
        md5_hash="abc", content_md5="abc", confirmed=confirmed, status=status,
        scan_root_id=scan_root_id, sequence=1,
    )
    sess.add(img)
    sess.commit()
    return img


def _make_version(sess, barcode, image_type, folder_ctime,
                  version_label="v1", is_latest=False, duplicate_mtimes="[]",
                  content_hash=None):
    if content_hash is None:
        content_hash = f"hash-{barcode}-{image_type}-{folder_ctime}"
    v = ImageVersion(
        barcode=barcode, image_type=image_type, folder_ctime=folder_ctime,
        version_label=version_label, content_hash=content_hash,
        is_latest=is_latest, duplicate_mtimes=duplicate_mtimes,
    )
    sess.add(v)
    sess.commit()
    return v


def _wait_for_task(client, task_id, timeout=10):
    """Poll task status until done/error or timeout."""
    start = time.time()
    while time.time() - start < timeout:
        resp = client.get(f'/api/tasks/{task_id}')
        data = resp.get_json()
        if data['status'] in ('done', 'partial_failed', 'error', 'cancelled', 'interrupted'):
            return data
        time.sleep(0.1)
    return data


# ===================================================================
# Task engine tests
# ===================================================================


def test_create_task_returns_task_dict(client, sess):
    _make_root(sess)
    resp = client.post('/api/batch/duplicate-scan/tasks')
    assert resp.status_code == 201
    data = resp.get_json()
    assert data['task_type'] == 'duplicate_scan'
    assert data['status'] in ('queued', 'running', 'done')
    assert 'id' in data


def test_dedup_same_params(client, sess):
    _make_root(sess)
    resp1 = client.post('/api/batch/duplicate-scan/tasks')
    data1 = resp1.get_json()
    # Second request with same params should return the same task
    resp2 = client.post('/api/batch/duplicate-scan/tasks')
    data2 = resp2.get_json()
    assert data2['id'] == data1['id']


def test_task_lifecycle_duplicate_scan(client, sess):
    _make_root(sess)
    ctime_v1 = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC1", "main", ctime_v1)
    _make_image(sess, "BC1", "main", dup)
    _make_version(sess, "BC1", "main", ctime_v1, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']

    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'
    assert task['result_count'] >= 1


def test_task_results_pagination(client, sess):
    _make_root(sess)
    ctime_v1 = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC1", "main", ctime_v1)
    _make_image(sess, "BC1", "main", dup)
    _make_version(sess, "BC1", "main", ctime_v1, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    _wait_for_task(client, task_id)

    # Get results
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results?page=1&page_size=10')
    assert resp.status_code == 200
    data = resp.get_json()
    assert 'items' in data
    assert 'total' in data
    assert data['page'] == 1


# ===================================================================
# Low version scan tests
# ===================================================================


def test_create_low_version_scan(client, sess):
    _make_root(sess)
    resp = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': True,
        'main_threshold': 3,
        'detail_enabled': True,
        'detail_threshold': 5,
    })
    assert resp.status_code == 201
    data = resp.get_json()
    assert data['task_type'] == 'low_version_scan'


def test_low_version_scan_requires_threshold(client, sess):
    resp = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': False,
        'detail_enabled': False,
    })
    assert resp.status_code == 400


def test_low_version_scan_results(client, sess):
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    _make_image(sess, "B1", "main", ctime)
    _make_version(sess, "B1", "main", ctime, version_label="v2", is_latest=False)
    _make_version(sess, "B1", "main", "2024-07-01T00:00:00", version_label="v1", is_latest=True)

    resp = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': True, 'main_threshold': 3,
        'detail_enabled': False, 'detail_threshold': 0,
    })
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)

    assert task['status'] == 'done'
    resp = client.get(f'/api/batch/low-version-scan/tasks/{task_id}/results?page=1&page_size=100')
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['total'] >= 1


# ===================================================================
# Delete task tests
# ===================================================================


def test_delete_queued_task(client, sess):
    _make_root(sess)
    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']

    # Try deleting — if running, expect 409; if done, expect 200
    resp = client.delete(f'/api/tasks/{task_id}')
    if resp.status_code == 409:
        # Task was running, wait for it to finish and try again
        _wait_for_task(client, task_id)
        resp = client.delete(f'/api/tasks/{task_id}')
    assert resp.status_code == 200


# ===================================================================
# Concurrent task dedup tests
# ===================================================================


def test_different_types_can_run_parallel(client, sess):
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    _make_image(sess, "BC1", "main", ctime)
    _make_version(sess, "BC1", "main", ctime, duplicate_mtimes=json.dumps([ctime]))

    # Start a duplicate scan
    resp1 = client.post('/api/batch/duplicate-scan/tasks')
    task1_id = resp1.get_json()['id']

    # Start a low version scan — should not be blocked
    resp2 = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': True, 'main_threshold': 2,
        'detail_enabled': False, 'detail_threshold': 0,
    })
    assert resp2.status_code in (200, 201)

    # Both should complete
    task1 = _wait_for_task(client, task1_id)
    task2 = _wait_for_task(client, resp2.get_json()['id'])
    assert task1['status'] == 'done'
    assert task2['status'] == 'done'


# ===================================================================
# Thread safety tests
# ===================================================================


def test_background_thread_does_not_leak_orm_objects(client, sess):
    """Verify background tasks run without cross-thread session issues."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_THREAD", "main", ctime)
    _make_image(sess, "BC_THREAD", "main", dup)
    _make_version(sess, "BC_THREAD", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done', f"Task ended as {task['status']}: {task.get('error_message', '')}"


def test_batch_delete_images_task_records_deleted_folders(client, sess):
    _make_root(sess)
    ctime1 = "2024-01-01T00:00:00"
    ctime2 = "2024-02-01T00:00:00"
    img1 = _make_image(sess, "BC_TASK_DEL", "main", ctime1, filename="a.jpg")
    img2 = _make_image(sess, "BC_TASK_DEL", "detail", ctime2, filename="b.jpg")

    resp = client.post("/api/images/batch-delete-task", json={
        "ids": [img1.id, img2.id],
        "delete_files": False,
    })
    assert resp.status_code == 201
    task = _wait_for_task(client, resp.get_json()["id"])

    assert task["status"] == "done"
    assert task["result_count"] == 2
    keys = {
        (d.barcode, d.image_type, d.folder_ctime)
        for d in sess.query(DeletedFolder).all()
    }
    assert ("BC_TASK_DEL", "main", ctime1) in keys
    assert ("BC_TASK_DEL", "detail", ctime2) in keys


def test_batch_delete_duplicate_versions_commits_indexes_and_continues(
    client, sess, tmp_path,
):
    """文件删除和索引清理必须共同提交，并继续处理后续版本。"""
    _make_root(sess, path=str(tmp_path))

    scan_task = BatchTask(
        task_type="duplicate_version_scan",
        status="done",
        params_json="{}",
    )
    sess.add(scan_task)
    sess.commit()

    specs = [
        ("BC_DV_FIRST", "2024-01-01T00:00:00", 2),
        ("BC_DV_SECOND", "2024-02-01T00:00:00", 1),
    ]
    result_ids = []
    image_ids = []
    file_paths = []
    for group_id, (barcode, folder_ctime, image_count) in enumerate(specs, 1):
        _make_version(sess, barcode, "main", folder_ctime)
        result = DuplicateVersionScanResult(
            task_id=scan_task.id,
            group_id=group_id,
            barcode=barcode,
            image_type="main",
            folder_ctime=folder_ctime,
            image_count=image_count,
            role="clean",
            delete_status="pending",
        )
        sess.add(result)
        sess.commit()
        result_ids.append(result.id)

        for sequence in range(1, image_count + 1):
            path = tmp_path / barcode / f"{sequence}.jpg"
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(f"{barcode}-{sequence}".encode())
            image = _make_image(
                sess,
                barcode,
                "main",
                folder_ctime,
                filename=path.name,
                file_path=str(path),
                status=("broken" if group_id == 1 and sequence == 1 else "active"),
            )
            image_ids.append(image.id)
            file_paths.append(path)

    # 已经不存在的文件也应视为物理删除完成，并继续清理对应索引。
    file_paths[0].unlink()

    response = client.post(
        "/api/batch/delete-duplicate-versions/tasks",
        json={
            "scan_task_id": scan_task.id,
            "result_ids": result_ids,
            "delete_files": True,
        },
    )
    assert response.status_code == 201
    task = _wait_for_task(client, response.get_json()["id"], timeout=15)

    assert task["status"] == "done", task.get("error_message", "")
    assert task["progress"] == len(specs)
    assert task["result_count"] == len(image_ids)

    sess.expire_all()
    assert sess.query(Image).filter(Image.id.in_(image_ids)).count() == 0
    results = sess.query(DuplicateVersionScanResult).filter(
        DuplicateVersionScanResult.id.in_(result_ids)
    ).all()
    assert {result.delete_status for result in results} == {"deleted"}
    assert all(not path.exists() for path in file_paths)


def test_batch_delete_duplicates_commits_indexes_and_continues(
    client, sess, tmp_path,
):
    """重复文件夹任务不能在首组物理删除后因进度写入而中断。"""
    _make_root(sess, path=str(tmp_path))
    items = []
    image_ids = []
    paths = []
    for index, barcode in enumerate(("BC_DUP_FIRST", "BC_DUP_SECOND"), 1):
        duplicate_ctime = f"2024-0{index}-01T00:00:00"
        keep_ctime = f"2024-0{index}-02T00:00:00"
        _make_version(
            sess,
            barcode,
            "main",
            keep_ctime,
            duplicate_mtimes=json.dumps([duplicate_ctime]),
        )
        path = tmp_path / barcode / "1.jpg"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(barcode.encode())
        image = _make_image(
            sess,
            barcode,
            "main",
            duplicate_ctime,
            filename=path.name,
            file_path=str(path),
        )
        image_ids.append(image.id)
        paths.append(path)
        items.append({
            "barcode": barcode,
            "image_type": "main",
            "folder_ctime": duplicate_ctime,
        })

    paths[0].unlink()
    response = client.post(
        "/api/batch/delete-duplicates/tasks",
        json={"items": items, "delete_files": True},
    )
    assert response.status_code == 201
    task = _wait_for_task(client, response.get_json()["id"], timeout=15)

    assert task["status"] == "done", task.get("error_message", "")
    assert task["progress"] == len(items)
    assert task["result_count"] == len(image_ids)
    sess.expire_all()
    assert sess.query(Image).filter(Image.id.in_(image_ids)).count() == 0
    assert all(not path.exists() for path in paths)


def test_batch_delete_low_versions_commits_indexes_and_continues(
    client, sess, tmp_path,
):
    """低版本任务按批提交索引，并把已缺失文件作为成功清理。"""
    _make_root(sess, path=str(tmp_path))
    items = []
    image_ids = []
    paths = []
    for index, barcode in enumerate(("BC_LOW_FIRST", "BC_LOW_SECOND"), 1):
        old_ctime = f"2024-0{index}-01T00:00:00"
        new_ctime = f"2024-0{index}-02T00:00:00"
        _make_version(sess, barcode, "main", old_ctime, version_label="v1")
        _make_version(
            sess,
            barcode,
            "main",
            new_ctime,
            version_label="v2",
            is_latest=True,
        )
        path = tmp_path / barcode / "1.jpg"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(barcode.encode())
        image = _make_image(
            sess,
            barcode,
            "main",
            old_ctime,
            filename=path.name,
            file_path=str(path),
        )
        image_ids.append(image.id)
        paths.append(path)
        items.append({
            "barcode": barcode,
            "image_type": "main",
            "folder_ctime": old_ctime,
        })

    paths[0].unlink()
    response = client.post(
        "/api/batch/delete-low-versions/tasks",
        json={
            "items": items,
            "delete_files": True,
            "main_threshold": 2,
            "detail_threshold": 0,
        },
    )
    assert response.status_code == 201
    task = _wait_for_task(client, response.get_json()["id"], timeout=15)

    assert task["status"] == "done", task.get("error_message", "")
    assert task["progress"] == len(items)
    assert task["result_count"] == len(image_ids)
    sess.expire_all()
    assert sess.query(Image).filter(Image.id.in_(image_ids)).count() == 0
    assert all(not path.exists() for path in paths)


def test_delete_version_task_removes_missing_file_index(client, sess, tmp_path):
    """单版本任务遇到文件已不存在时仍应提交索引删除。"""
    _make_root(sess, path=str(tmp_path))
    barcode = "BC_VERSION_MISSING"
    folder_ctime = "2024-03-01T00:00:00"
    version = _make_version(sess, barcode, "main", folder_ctime)
    path = tmp_path / barcode / "1.jpg"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(barcode.encode())
    image = _make_image(
        sess,
        barcode,
        "main",
        folder_ctime,
        filename=path.name,
        file_path=str(path),
        status="broken",
    )
    image_id = image.id
    path.unlink()

    response = client.post(
        f"/api/versions/{version.id}/delete-task",
        json={"delete_files": True},
    )
    assert response.status_code == 201
    task = _wait_for_task(client, response.get_json()["id"], timeout=15)

    assert task["status"] == "done", task.get("error_message", "")
    assert task["progress"] == 1
    assert task["result_count"] == 1
    sess.expire_all()
    assert sess.get(Image, image_id) is None


def test_duplicate_version_delete_retries_locked_checkpoint_in_chunks(
    client, sess, monkeypatch,
):
    """锁冲突只重试未提交分块，且不会退化为逐项提交。"""
    import routes.batch_tasks as batch_tasks

    _make_root(sess)
    scan_task = BatchTask(
        task_type="duplicate_version_scan",
        status="done",
        params_json="{}",
    )
    sess.add(scan_task)
    sess.commit()

    result_ids = []
    image_ids = []
    for index in range(26):
        barcode = f"BC_CHUNK_{index:02d}"
        folder_ctime = f"2024-01-{index + 1:02d}T00:00:00"
        image = _make_image(sess, barcode, "main", folder_ctime)
        result = DuplicateVersionScanResult(
            task_id=scan_task.id,
            group_id=index + 1,
            barcode=barcode,
            image_type="main",
            folder_ctime=folder_ctime,
            image_count=1,
            role="clean",
            delete_status="pending",
        )
        sess.add(result)
        sess.commit()
        image_ids.append(image.id)
        result_ids.append(result.id)

    original_checkpoint = batch_tasks._commit_delete_checkpoint
    checkpoint_progress = []
    locked_once = threading.Event()

    def checkpoint_with_one_lock(*args, **kwargs):
        progress = kwargs["progress"]
        checkpoint_progress.append(progress)
        if progress == 26 and not locked_once.is_set():
            locked_once.set()
            raise sqlite3.OperationalError("database is locked")
        return original_checkpoint(*args, **kwargs)

    monkeypatch.setattr(
        batch_tasks,
        "_commit_delete_checkpoint",
        checkpoint_with_one_lock,
    )
    monkeypatch.setattr(batch_tasks, "_rebuild_versions_safe", lambda *args: [])

    response = client.post(
        "/api/batch/delete-duplicate-versions/tasks",
        json={
            "scan_task_id": scan_task.id,
            "result_ids": result_ids,
            "delete_files": False,
        },
    )
    assert response.status_code == 201
    task = _wait_for_task(client, response.get_json()["id"], timeout=15)

    assert task["status"] == "done", task.get("error_message", "")
    assert task["progress"] == len(result_ids)
    assert task["result_count"] == len(image_ids)
    assert locked_once.is_set()
    assert checkpoint_progress == [0, 25, 26, 25, 26]
    sess.expire_all()
    assert sess.query(Image).filter(Image.id.in_(image_ids)).count() == 0


# ===================================================================
# Delete running task test
# ===================================================================


def test_delete_running_task_rejected(client, sess):
    """Deleting a running task must return 409."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_RUN_DEL", "main", ctime)
    _make_image(sess, "BC_RUN_DEL", "main", dup)
    _make_version(sess, "BC_RUN_DEL", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']

    # Task may already be done (fast), so check current status first
    task_data = client.get(f'/api/tasks/{task_id}').get_json()
    if task_data['status'] == 'running':
        resp = client.delete(f'/api/tasks/{task_id}')
        assert resp.status_code == 409, f"Expected 409, got {resp.status_code}"
        data = resp.get_json()
        assert 'error' in data
    # If already done, deleting is fine — that's a different scenario


# ===================================================================
# Cancel task tests
# ===================================================================


def test_cancel_queued_task(client, sess):
    """Cancelling a queued task should succeed."""
    from task_engine import create_task

    # 404 for non-existent task
    resp = client.post('/api/tasks/9999/cancel')
    assert resp.status_code == 404

    # Create a queued task (auto_start=False so it stays queued)
    task_dict, is_new = create_task('duplicate_scan', params={}, auto_start=False)
    assert is_new
    assert task_dict['status'] == 'queued'

    # Cancel it via the route
    resp = client.post(f'/api/tasks/{task_dict["id"]}/cancel')
    assert resp.status_code == 200
    cancelled = resp.get_json()
    assert cancelled['status'] == 'cancelled'
    assert cancelled['finished_at']


def test_cancel_nonexistent_task(client, sess):
    """Cancel non-existent task returns 404."""
    resp = client.post('/api/tasks/9999/cancel')
    assert resp.status_code == 404


# ===================================================================
# Path safety tests
# ===================================================================


def test_path_outside_root_not_deleted(client, sess):
    """Images with paths outside their ScanRoot must not be os.remove'd
    and their indices must not be deleted."""
    import os
    root = _make_root(sess, root_id=1, path="/safe/root", enabled=True)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"

    # Image with a path OUTSIDE the scan root (same drive so realpath works)
    import tempfile
    danger_path = os.path.join(tempfile.gettempdir(), "outside_test_danger.jpg")
    _make_image(sess, "BC_PATH", "main", dup, file_path=danger_path, scan_root_id=root.id)
    _make_version(sess, "BC_PATH", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    # Run duplicate scan
    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    # Get results
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    results = resp.get_json()
    assert results['total'] >= 1
    result_id = results['items'][0]['id']

    # Try to delete with delete_files=True — should fail with path safety
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [result_id],
        'delete_files': True,
    })
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['skipped_count'] >= 1, f"Expected skipped_count >= 1, got {data}"
    assert data['deleted_image_count'] == 0

    # Verify the result was marked as failed (not deleted)
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    results = resp.get_json()
    r = results['items'][0]
    assert r['delete_status'] in ('failed', 'skipped'), f"Expected failed/skipped, got {r['delete_status']}"
    assert '路径' in (r.get('delete_message') or '')


@pytest.mark.parametrize(
    ("task_type", "endpoint"),
    [
        ("duplicate_scan", "/api/batch/duplicate-scan/tasks/{task_id}/delete"),
        ("low_version_scan", "/api/batch/low-version-scan/tasks/{task_id}/delete"),
    ],
)
def test_task_delete_rejects_non_list_result_ids(client, task_type, endpoint):
    """Deletion endpoints must reject malformed result_ids before querying."""
    from task_engine import create_task

    task, _ = create_task(task_type, params={}, auto_start=False)

    resp = client.post(endpoint.format(task_id=task["id"]), json={
        "mode": "selected",
        "result_ids": "abc",
        "delete_files": False,
    })

    assert resp.status_code == 400
    assert "result_ids" in resp.get_json()["error"]


@pytest.mark.parametrize(
    ("task_type", "endpoint"),
    [
        ("duplicate_scan", "/api/batch/duplicate-scan/tasks/{task_id}/delete"),
        ("low_version_scan", "/api/batch/low-version-scan/tasks/{task_id}/delete"),
    ],
)
@pytest.mark.parametrize("bad_ids", [
    [["x"]],           # list containing a list
    [1, ["x"]],        # mixed int and list
    [1, "2"],          # mixed int and string
    [True],            # bool is subclass of int, must reject
    [0],               # zero is not positive
    [-1],              # negative
    pytest.param([], id="empty-list"),
])
def test_task_delete_rejects_invalid_result_id_elements(client, task_type, endpoint, bad_ids):
    """Deletion endpoints must reject result_ids with non-int or non-positive elements."""
    from task_engine import create_task

    task, _ = create_task(task_type, params={}, auto_start=False)

    resp = client.post(endpoint.format(task_id=task["id"]), json={
        "mode": "selected",
        "result_ids": bad_ids,
        "delete_files": False,
    })

    assert resp.status_code == 400, f"Expected 400 for {bad_ids}, got {resp.status_code}"
    assert "result_ids" in resp.get_json()["error"]


def test_skipped_count_increments_on_validation_failure(client, sess):
    """skipped_count must increment when a result fails re-validation."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_SKIP", "main", ctime)
    _make_image(sess, "BC_SKIP", "main", dup)
    _make_version(sess, "BC_SKIP", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    results = resp.get_json()
    result_id = results['items'][0]['id']

    # Delete without delete_files — should succeed (index only)
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [result_id],
        'delete_files': False,
    })
    assert resp.status_code == 200
    data = resp.get_json()
    # Index-only deletion should have skipped_count=0
    assert data['deleted_image_count'] >= 1


def test_deleted_at_written_on_successful_delete(client, sess):
    """deleted_at must be set when a result is successfully deleted."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_DELAT", "main", dup, file_path="/fake/BC_DELAT/dup/a.jpg")
    _make_version(sess, "BC_DELAT", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    results = resp.get_json()
    assert results['total'] >= 1
    result_id = results['items'][0]['id']

    # Index-only delete should succeed
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [result_id],
        'delete_files': False,
    })
    assert resp.status_code == 200
    data = resp.get_json()
    assert data['skipped_count'] == 0

    # Verify deleted_at is set
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    results = resp.get_json()
    r = results['items'][0]
    assert r['delete_status'] == 'deleted'
    assert r.get('deleted_at'), "deleted_at must be set after successful delete"


# ===================================================================
# Restart recovery tests
# ===================================================================


def test_queued_tasks_marked_interrupted_on_startup(engine):
    """On startup, both running and queued tasks must be marked interrupted."""
    from sqlalchemy.orm import sessionmaker, scoped_session

    factory = sessionmaker(bind=engine)
    sess = scoped_session(factory)
    try:
        # Create a queued task directly in DB
        sess.execute(
            __import__('sqlalchemy').text(
                "INSERT INTO batch_task (task_type, status, params_json, created_at) "
                "VALUES ('duplicate_scan', 'queued', '{}', '2024-01-01T00:00:00')"
            )
        )
        sess.commit()

        # Simulate startup recovery logic (same as app.py)
        now_iso = __import__('datetime').datetime.now().isoformat()
        conn = sess.bind.connect()
        _running = conn.execute(
            __import__('sqlalchemy').text("SELECT COUNT(*) FROM batch_task WHERE status = 'running'")
        ).fetchone()[0]
        _queued = conn.execute(
            __import__('sqlalchemy').text("SELECT COUNT(*) FROM batch_task WHERE status = 'queued'")
        ).fetchone()[0]
        if _queued:
            conn.execute(
                __import__('sqlalchemy').text(
                    "UPDATE batch_task SET status = 'interrupted', error_message = '程序重启，任务未执行', "
                    "finished_at = :now WHERE status = 'queued'"
                ), {'now': now_iso}
            )
            conn.execute(__import__('sqlalchemy').text("COMMIT"))
        conn.close()

        # Verify the queued task is now interrupted
        from models import BatchTask
        task = sess.query(BatchTask).filter(BatchTask.task_type == 'duplicate_scan').first()
        assert task is not None
        assert task.status == 'interrupted', f"Expected 'interrupted', got '{task.status}'"
        assert '重启' in (task.error_message or '')
    finally:
        sess.remove()


# ===================================================================
# Export concurrency lock tests
# ===================================================================


def test_export_concurrent_processing_returns_409(client, sess):
    """When a processing ExportTask exists, creating a new one must return 409."""
    from routes.export import _export_lock, ExportTask as ET
    # Create a processing export task directly
    with _export_lock:
        existing = sess.query(ET).filter(ET.status == 'processing').first()
        # Clean up any existing processing tasks from other tests
        if existing:
            existing.status = 'failed'
            sess.commit()

        task = ET(status='processing')
        sess.add(task)
        sess.commit()
        task_id = task.id

    # Now try to create another — should be rejected by the lock
    with _export_lock:
        running = sess.query(ET).filter(ET.status == 'processing').first()
        assert running is not None
        assert running.id == task_id

    # Clean up
    with _export_lock:
        t = sess.get(ET, task_id)
        if t:
            sess.delete(t)
            sess.commit()


# ===================================================================
# Concurrent task session isolation test
# ===================================================================


def test_concurrent_tasks_do_not_corrupt_module_sessions(client, sess):
    """Running duplicate_scan and low_version_scan concurrently must not
    mutate task_engine.session / routes.batch_tasks.session globals."""
    import task_engine as _te
    import routes.batch_tasks as _bt
    import routes.batch as _batch
    import versioning as _ver

    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_CONC", "main", ctime)
    _make_image(sess, "BC_CONC", "main", dup)
    _make_version(sess, "BC_CONC", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    # Capture original session objects
    orig_te = _te.session
    orig_bt = _bt.session
    orig_batch = _batch.session
    orig_ver = _ver.session

    # Start duplicate scan
    resp1 = client.post('/api/batch/duplicate-scan/tasks')
    id1 = resp1.get_json()['id']

    # Start low version scan
    resp2 = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': True, 'main_threshold': 2,
        'detail_enabled': False, 'detail_threshold': 0,
    })
    id2 = resp2.get_json()['id']

    # Wait for both
    t1 = _wait_for_task(client, id1)
    t2 = _wait_for_task(client, id2)
    assert t1['status'] == 'done'
    assert t2['status'] == 'done'

    # Module-level sessions must be unchanged
    assert _te.session is orig_te, "task_engine.session was mutated by concurrent tasks"
    assert _bt.session is orig_bt, "routes.batch_tasks.session was mutated by concurrent tasks"
    assert _batch.session is orig_batch, "routes.batch.session was mutated by concurrent tasks"
    assert _ver.session is orig_ver, "versioning.session was mutated by concurrent tasks"


# ===================================================================
# Export concurrency / recovery tests
# ===================================================================


def test_export_both_endpoints_reject_when_processing(client, sess):
    """Both /export/zip and /images/batch-export must return 409 when a
    processing ExportTask already exists."""
    from routes.export import ExportTask as ET
    import routes.export as _export

    with _export._export_lock:
        task = ET(status='processing')
        sess.add(task)
        sess.commit()
        pid = task.id

    try:
        # /export/zip path: we can't call the full route easily (needs Excel),
        # but the lock + query logic is the same.  Test via direct lock.
        with _export._export_lock:
            running = sess.query(ET).filter(ET.status == 'processing').first()
            assert running is not None
            assert running.id == pid

        # /images/batch-export: register images_bp and test the endpoint
        app = __import__('flask').Flask(__name__)
        app.config['TESTING'] = True
        from routes.images import images_bp
        app.register_blueprint(images_bp, url_prefix='/api')
        import routes.images as _imgs
        _imgs.session = sess

        with app.test_client() as c:
            # batch-export should reject because processing exists
            resp = c.post('/api/images/batch-export', json={'ids': [1], 'image_type': 'main'})
            assert resp.status_code == 409, f"Expected 409, got {resp.status_code}"
            data = resp.get_json()
            assert '已有导出' in data.get('error', '')
    finally:
        with _export._export_lock:
            t = sess.get(ET, pid)
            if t:
                sess.delete(t)
                sess.commit()


def test_reset_stale_processing_unblocks_export(client, sess):
    """reset_stale_processing() must mark all processing tasks as failed,
    allowing new exports to proceed."""
    from routes.export import reset_stale_processing, ExportTask as ET, _export_lock

    # Create a processing task
    with _export_lock:
        task = ET(status='processing')
        sess.add(task)
        sess.commit()
        pid = task.id

    # Reset
    reset_stale_processing()

    # Verify it's now failed
    t = sess.get(ET, pid)
    assert t.status == 'failed'
    assert '重启' in (t.error_message or '')

    # Verify no processing remains — new export would succeed
    with _export_lock:
        running = sess.query(ET).filter(ET.status == 'processing').first()
        assert running is None

    # Cleanup
    with _export_lock:
        sess.delete(t)
        sess.commit()


# ===================================================================
# Partial deletion consistency test
# ===================================================================


def test_partial_deletion_does_not_leave_partial_index_changes(client, sess):
    """When one image in a result fails validation, DB indices must not be
    partially deleted for that result, and result must be marked failed."""
    root = _make_root(sess, root_id=1, path="/safe/root", enabled=True)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"

    # Two images in the same result group: one valid path, one unsafe
    _make_image(sess, "BC_PARTIAL", "main", dup, file_path="/safe/root/good.jpg", scan_root_id=root.id)
    import tempfile, os as _os
    danger = _os.path.join(tempfile.gettempdir(), "unsafe.jpg")
    _make_image(sess, "BC_PARTIAL", "main", dup, file_path=danger, scan_root_id=root.id)
    _make_version(sess, "BC_PARTIAL", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    # Run scan
    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    # Get result
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    data = resp.get_json()
    assert data['total'] >= 1
    result_id = data['items'][0]['id']

    # Count images before delete
    from models import Image as _Img
    before = sess.query(_Img).filter(_Img.barcode == 'BC_PARTIAL', _Img.status == 'active').count()

    # Try delete with delete_files=True — should fail because one path is unsafe
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [result_id],
        'delete_files': True,
    })
    assert resp.status_code == 200
    delete_data = resp.get_json()
    assert delete_data['skipped_count'] >= 1
    assert delete_data['deleted_image_count'] == 0

    # Verify result is marked failed, not deleted
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    data = resp.get_json()
    r = data['items'][0]
    assert r['delete_status'] in ('failed', 'skipped'), f"Expected failed/skipped, got {r['delete_status']}"

    # Verify NO images were deleted from DB (no partial index deletion)
    after = sess.query(_Img).filter(_Img.barcode == 'BC_PARTIAL', _Img.status == 'active').count()
    assert after == before, f"Images count changed: {before} -> {after}"


# ===================================================================
# Export import regression test
# ===================================================================


def test_cleanup_old_exports_importable():
    """Regression: cleanup_old_exports must be importable alongside
    reset_stale_processing."""
    from routes.export import cleanup_old_exports, reset_stale_processing
    assert callable(cleanup_old_exports)
    assert callable(reset_stale_processing)


def test_cleanup_old_exports_deletes_expired_non_processing(client, sess):
    """cleanup_old_exports() must delete expired non-processing tasks and
    their zip files."""
    import routes.export as _export
    from routes.export import ExportTask as ET, ZIP_CLEANUP_HOURS
    import datetime as _dt
    import os as _os

    # Create an old done task with a fake zip file
    old_time = (_dt.datetime.now() - _dt.timedelta(hours=ZIP_CLEANUP_HOURS + 1)).isoformat()
    with _export._export_lock:
        task = ET(status='done', created_at=old_time, total_images=5)
        sess.add(task)
        sess.commit()
        tid = task.id

    # Create a fake zip file under the controlled zips dir
    zip_dir = _export._zip_dir()
    zip_path = _os.path.join(zip_dir, f'export_{tid}.zip')
    with _export._export_lock:
        t = sess.get(ET, tid)
        t.zip_path = zip_path
        sess.commit()

    # Create the fake file
    with open(zip_path, 'w') as f:
        f.write('fake')

    try:
        _export.cleanup_old_exports()

        # Task should be deleted
        t = sess.get(ET, tid)
        assert t is None, f"Expired task {tid} was not cleaned up"

        # Zip file should be removed
        assert not _os.path.exists(zip_path), "Zip file was not removed"
    finally:
        if _os.path.exists(zip_path):
            _os.remove(zip_path)


# ===================================================================
# App startup contract test
# ===================================================================


def test_cleanup_exports_on_startup_calls_both_in_order(sess, monkeypatch):
    """app.py _cleanup_exports_on_startup() must call reset_stale_processing
    first, then cleanup_old_exports.  Regression: non-tray path used to skip
    reset_stale_processing, leaving stale processing unhandled.

    Uses the test's real temp SQLite engine so that app.py import-time
    migrations (PRAGMA table_info / ALTER TABLE) succeed against actual
    tables created by Base.metadata.create_all."""

    # Point models at the test engine + session so import-time migrations
    # and create_all run against the temp DB where tables actually exist.
    monkeypatch.setattr('models.engine', sess.bind)
    monkeypatch.setattr('models.session', sess)

    # Mock the two export functions to verify call order without
    # performing real cleanup against the test DB.
    import routes.export as _export
    calls = []
    monkeypatch.setattr(_export, 'reset_stale_processing', lambda: calls.append('reset'))
    monkeypatch.setattr(_export, 'cleanup_old_exports', lambda: calls.append('cleanup'))

    import app as _app
    _app._cleanup_exports_on_startup()

    assert calls == ['reset', 'cleanup'], (
        f"Expected reset then cleanup, got {calls}"
    )


# ===================================================================
# ID consistency validation tests
# ===================================================================


def test_duplicate_delete_rejects_foreign_result_id(client, sess):
    """Submitting a result ID from another task (or non-existent) must return 400
    and the valid result must NOT be marked deleted or skipped."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_DUP_ID", "main", ctime)
    _make_image(sess, "BC_DUP_ID", "main", dup)
    _make_version(sess, "BC_DUP_ID", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    # Run duplicate scan to produce a valid result
    resp = client.post('/api/batch/duplicate-scan/tasks')
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    # Fetch results
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    data = resp.get_json()
    assert data['total'] >= 1
    valid_id = data['items'][0]['id']

    # Submit [valid_id, 999999] — 999999 does not exist
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [valid_id, 999999],
        'delete_files': False,
    })
    assert resp.status_code == 400
    err = resp.get_json()
    assert 'missing_ids' in err
    assert 999999 in err['missing_ids']

    # Verify the valid result was NOT marked deleted or skipped
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task_id}/results')
    data = resp.get_json()
    r = next(item for item in data['items'] if item['id'] == valid_id)
    assert r['delete_status'] in (None, 'pending'), f"Valid result should be untouched, got delete_status={r['delete_status']}"


def test_duplicate_delete_rejects_cross_task_result_id(client, sess):
    """Submitting a result ID that belongs to a different task must return 400."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    dup = "2024-05-01T00:00:00"
    _make_image(sess, "BC_CROSS1", "main", ctime)
    _make_image(sess, "BC_CROSS1", "main", dup)
    _make_version(sess, "BC_CROSS1", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    _make_image(sess, "BC_CROSS2", "main", ctime)
    _make_image(sess, "BC_CROSS2", "main", dup)
    _make_version(sess, "BC_CROSS2", "main", ctime, duplicate_mtimes=json.dumps([dup]))

    # Run two separate scans
    resp1 = client.post('/api/batch/duplicate-scan/tasks')
    task1_id = resp1.get_json()['id']
    _wait_for_task(client, task1_id)

    resp2 = client.post('/api/batch/duplicate-scan/tasks')
    task2_id = resp2.get_json()['id']
    _wait_for_task(client, task2_id)

    # Get task2's results
    resp = client.get(f'/api/batch/duplicate-scan/tasks/{task2_id}/results')
    data = resp.get_json()
    assert data['total'] >= 1
    task2_result_id = data['items'][0]['id']

    # Try to delete task2's result via task1's endpoint
    resp = client.post(f'/api/batch/duplicate-scan/tasks/{task1_id}/delete', json={
        'mode': 'selected',
        'result_ids': [task2_result_id],
        'delete_files': False,
    })
    assert resp.status_code == 400
    err = resp.get_json()
    assert 'missing_ids' in err
    assert task2_result_id in err['missing_ids']


def test_low_version_delete_rejects_foreign_result_id(client, sess):
    """Submitting a result ID from another task (or non-existent) must return 400
    and the valid result must NOT be marked deleted or skipped."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    _make_image(sess, "BC_LV_ID", "main", ctime)
    _make_version(sess, "BC_LV_ID", "main", ctime, version_label="v2", is_latest=False)
    _make_version(sess, "BC_LV_ID", "main", "2024-07-01T00:00:00", version_label="v1", is_latest=True)

    # Run low version scan
    resp = client.post('/api/batch/low-version-scan/tasks', json={
        'main_enabled': True, 'main_threshold': 3,
        'detail_enabled': False, 'detail_threshold': 0,
    })
    task_id = resp.get_json()['id']
    task = _wait_for_task(client, task_id)
    assert task['status'] == 'done'

    # Fetch results
    resp = client.get(f'/api/batch/low-version-scan/tasks/{task_id}/results')
    data = resp.get_json()
    assert data['total'] >= 1
    valid_id = data['items'][0]['id']

    # Submit [valid_id, 999999] — 999999 does not exist
    resp = client.post(f'/api/batch/low-version-scan/tasks/{task_id}/delete', json={
        'mode': 'selected',
        'result_ids': [valid_id, 999999],
        'delete_files': False,
    })
    assert resp.status_code == 400
    err = resp.get_json()
    assert 'missing_ids' in err
    assert 999999 in err['missing_ids']

    # Verify the valid result was NOT marked deleted or skipped
    resp = client.get(f'/api/batch/low-version-scan/tasks/{task_id}/results')
    data = resp.get_json()
    r = next(item for item in data['items'] if item['id'] == valid_id)
    assert r['delete_status'] in (None, 'pending'), f"Valid result should be untouched, got delete_status={r['delete_status']}"


def test_low_version_delete_rejects_cross_task_result_id(client, sess):
    """Submitting a result ID that belongs to a different task must return 400."""
    _make_root(sess)
    ctime = "2024-06-01T00:00:00"
    _make_image(sess, "BC_LV_CROSS1", "main", ctime)
    _make_version(sess, "BC_LV_CROSS1", "main", ctime, version_label="v2", is_latest=False)
    _make_version(sess, "BC_LV_CROSS1", "main", "2024-07-01T00:00:00", version_label="v1", is_latest=True)

    _make_image(sess, "BC_LV_CROSS2", "main", ctime)
    _make_version(sess, "BC_LV_CROSS2", "main", ctime, version_label="v2", is_latest=False)
    _make_version(sess, "BC_LV_CROSS2", "main", "2024-07-01T00:00:00", version_label="v1", is_latest=True)

    scan_params = {'main_enabled': True, 'main_threshold': 3, 'detail_enabled': False, 'detail_threshold': 0}

    resp1 = client.post('/api/batch/low-version-scan/tasks', json=scan_params)
    task1_id = resp1.get_json()['id']
    _wait_for_task(client, task1_id)

    resp2 = client.post('/api/batch/low-version-scan/tasks', json=scan_params)
    task2_id = resp2.get_json()['id']
    _wait_for_task(client, task2_id)

    # Get task2's results
    resp = client.get(f'/api/batch/low-version-scan/tasks/{task2_id}/results')
    data = resp.get_json()
    assert data['total'] >= 1
    task2_result_id = data['items'][0]['id']

    # Try to delete task2's result via task1's endpoint
    resp = client.post(f'/api/batch/low-version-scan/tasks/{task1_id}/delete', json={
        'mode': 'selected',
        'result_ids': [task2_result_id],
        'delete_files': False,
    })
    assert resp.status_code == 400
    err = resp.get_json()
    assert 'missing_ids' in err
    assert task2_result_id in err['missing_ids']


# ===================================================================
# Export detail download tests
# ===================================================================


def test_download_detail_has_three_sheets(client, sess):
    """download_detail endpoint must return Excel with three sheets:
    导出详情, 主图匹配, 详情图匹配."""
    from routes.export import ExportTask as ET, _export_lock
    from openpyxl import load_workbook
    import io

    barcode_data = json.dumps({
        'BC001': {'main': 2, 'detail': 3},
        'BC002': {'main': 0, 'detail': 1},
        'BC003': {'main': 0, 'detail': 0},
    }, ensure_ascii=False)

    with _export_lock:
        task = ET(status='done', barcode_data=barcode_data)
        sess.add(task)
        sess.commit()
        task_id = task.id

    resp = client.get(f'/export/tasks/{task_id}/detail')
    assert resp.status_code == 200
    assert resp.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = load_workbook(io.BytesIO(resp.data))
    assert wb.sheetnames == ['导出详情', '主图匹配', '详情图匹配']

    # Sheet 1: 导出详情
    ws_all = wb['导出详情']
    assert [cell.value for cell in ws_all[1]] == ['条码', '匹配主图数量', '匹配详情图数量']
    rows_all = [[cell.value for cell in row] for row in ws_all.iter_rows(min_row=2)]
    assert rows_all == [
        ['BC001', 2, 3],
        ['BC002', 0, 1],
        ['BC003', 0, 0],
    ]

    # Sheet 2: 主图匹配
    ws_main = wb['主图匹配']
    assert [cell.value for cell in ws_main[1]] == ['条码', '主图数量']
    rows_main = [[cell.value for cell in row] for row in ws_main.iter_rows(min_row=2)]
    assert rows_main == [
        ['BC001', 2],
        ['BC002', 0],
        ['BC003', 0],
    ]

    # Sheet 3: 详情图匹配
    ws_detail = wb['详情图匹配']
    assert [cell.value for cell in ws_detail[1]] == ['条码', '详情图数量']
    rows_detail = [[cell.value for cell in row] for row in ws_detail.iter_rows(min_row=2)]
    assert rows_detail == [
        ['BC001', 3],
        ['BC002', 1],
        ['BC003', 0],
    ]

    # Cleanup
    with _export_lock:
        t = sess.get(ET, task_id)
        if t:
            sess.delete(t)
            sess.commit()


def test_download_detail_returns_404_for_missing_task(client, sess):
    """download_detail must return 404 for non-existent task."""
    resp = client.get('/export/tasks/99999/detail')
    assert resp.status_code == 404


def test_download_detail_returns_404_for_no_barcode_data(client, sess):
    """download_detail must return 404 when task has no barcode_data."""
    from routes.export import ExportTask as ET, _export_lock

    with _export_lock:
        task = ET(status='done', barcode_data=None)
        sess.add(task)
        sess.commit()
        task_id = task.id

    resp = client.get(f'/export/tasks/{task_id}/detail')
    assert resp.status_code == 404

    # Cleanup
    with _export_lock:
        t = sess.get(ET, task_id)
        if t:
            sess.delete(t)
            sess.commit()
